"""
master_index.py — Builds (or refreshes) master_index.xlsx from all 11 brand files.

Run any time you want to sync the master:
    python master_index.py

Reads:  ./output/<Brand>_filaments_v3_N.xlsx     — highest N found, per brand
        ./output/Filament_Inventory_vN.xlsx      — highest N found
Writes: ./output/master_index_11_{N+1}.xlsx      — next N after the highest found

File discovery is version-aware (see template_v3.find_latest_version):
nothing here is a hardcoded filename, so this script always reads whatever
the current highest-versioned file is for each brand, regardless of how far
generate_all.py / generate_azurefilm.py / generate_voxelpla.py / build_inventory.py
have incremented since this script was last touched.

Sheet layout:
  1. Dashboard   — cross-brand summary stats and per-brand breakdown
  2. Catalog     — combined catalog from all 11 brands (read-only section + Brand column)
  3. Inventory   — combined inventory from all 11 brands + manual rows section
"""

import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

from template_v3 import find_latest_version, next_versioned_path

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

BRANDS = [
    "AzureFilm", "Bambu Lab", "SUNLU", "Polymaker", "eSUN", "Overture",
    "Hatchbox", "Prusament", "Creality Hyper", "MatterHackers", "VoxelPLA",
]


def _latest_brand_file(brand: str):
    """Return (path, filename) for the highest-versioned catalog workbook
    for `brand`, or (None, None) if none exists yet in OUTPUT_DIR."""
    safe = brand.replace(" ", "_")
    pattern = rf'^{re.escape(safe)}_filaments_v3_(\d+)\.xlsx$'
    _, fname = find_latest_version(OUTPUT_DIR, pattern)
    if not fname:
        return None, None
    return os.path.join(OUTPUT_DIR, fname), fname


def _latest_inventory_file():
    """Return (path, filename) for the highest-versioned inventory workbook,
    or (None, None) if none exists yet in OUTPUT_DIR."""
    pattern = r'^Filament_Inventory_v(\d+)\.xlsx$'
    _, fname = find_latest_version(OUTPUT_DIR, pattern)
    if not fname:
        return None, None
    return os.path.join(OUTPUT_DIR, fname), fname

# ── Palette (matches template_v3) ──────────────────────────────────────────────
P = {
    "hdr_bg":   "1F3864", "hdr_fg":   "FFFFFF",
    "sec_bg":   "2F5496", "sec_fg":   "FFFFFF",
    "alt":      "EEF3FB",
    "tier_s":   "FFD700", "tier_s_fg":"7F6000",
    "tier_a":   "92D050", "tier_a_fg":"276221",
    "tier_b":   "BDD7EE", "tier_b_fg":"1F3864",
    "tier_c":   "F4B8B8", "tier_c_fg":"9C0006",
    "ams_ok":   "C6EFCE", "ams_ok_fg":"276221",
    "ams_no":   "FFC7CE", "ams_no_fg":"9C0006",
    "ams_warn": "FFEB9C", "ams_warn_fg":"9C5700",
    "manual":   "FFFDE7",
    "reorder":  "FF0000", "reord_fg": "FFFFFF",
    "border":   "B8CCE4",
    "row_hl":   "F2F2F2",
}
FONT_NAME = "Arial"


# ── Style helpers ──────────────────────────────────────────────────────────────
def _argb(h):
    """Ensure full-opacity ARGB 8-char hex (openpyxl silently prepends 00=transparent)."""
    return h if len(h) == 8 else f"FF{h}"
def _fill(h):
    return PatternFill("solid", fgColor=_argb(h))
def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=_argb(color), italic=italic)
def _border():
    s = Side(style="thin", color=_argb(P["border"]))
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _tint_hex(hex6, factor=0.15):
    """Blend hex6 toward white (factor=0 -> white, factor=1 -> original)."""
    r = int(int(hex6[:2],16)*factor + 255*(1-factor))
    g = int(int(hex6[2:4],16)*factor + 255*(1-factor))
    b = int(int(hex6[4:6],16)*factor + 255*(1-factor))
    return f"{r:02X}{g:02X}{b:02X}"

def _sc(cell, bg=None, fg="000000", bold=False, size=10,
        align="left", border=True, italic=False):
    cell.font = _font(size=size, bold=bold, color=fg, italic=italic)
    if bg:  cell.fill = _fill(bg)
    cell.alignment = _align("center" if align == "center" else "left")
    if border: cell.border = _border()

def _header_row(ws, row, headers, bg=None, fg=None, height=30):
    bg, fg = bg or P["hdr_bg"], fg or P["hdr_fg"]
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        _sc(c, bg=bg, fg=fg, bold=True, align="center")
    ws.row_dimensions[row].height = height

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _merge(ws, r1, c1, r2, c2, val="", bg=None, fg="FFFFFF",
           bold=True, size=12, align="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=val)
    c.font = _font(size=size, bold=bold, color=fg)
    if bg: c.fill = _fill(bg)
    c.alignment = _align(align)

def _section(ws, row, text, ncols=10):
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    _sc(c, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, border=False)
    ws.row_dimensions[row].height = 18

def _cf_has(ws, col, r1, r2, val, bg, fg="000000"):
    rng = f"{col}{r1}:{col}{r2}"
    ds = DifferentialStyle(font=Font(name=FONT_NAME, color=_argb(fg)),
                           fill=PatternFill(bgColor=_argb(bg)))
    ws.conditional_formatting.add(
        rng,
        Rule(type="expression",
             formula=[f'NOT(ISERROR(SEARCH("{val}",${col}{r1})))'], dxf=ds),
    )

def _cf_eq(ws, col, r1, r2, val, bg, fg="000000"):
    rng = f"{col}{r1}:{col}{r2}"
    ds = DifferentialStyle(font=Font(name=FONT_NAME, color=_argb(fg)),
                           fill=PatternFill(bgColor=_argb(bg)))
    ws.conditional_formatting.add(
        rng,
        Rule(type="expression", formula=[f'${col}{r1}="{val}"'], dxf=ds),
    )


# ── Data extraction ────────────────────────────────────────────────────────────
def _read_sheet(path: str, sheet_name: str, header_row: int = 2,
                 required_col: str = None) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1)
        df = df.dropna(how="all")
        # Template rows in Inventory Tracker carry live Reorder Flag / Total
        # Value formulas (cached result 0) even when the row is otherwise
        # blank, so dropna(how="all") never removes them. Where a required
        # identifying column is given (e.g. SKU), drop rows that lack it —
        # that's the only column guaranteed blank on a phantom template row.
        if required_col and required_col in df.columns:
            df = df.dropna(subset=[required_col])
        return df
    except Exception:
        return pd.DataFrame()


def _load_all_data():
    catalog_frames = []

    for brand in BRANDS:
        path, filename = _latest_brand_file(brand)
        if not path:
            safe = brand.replace(" ", "_")
            print(f"  ⚠  Not found, skipping: {brand} "
                  f"(no {safe}_filaments_v3_N.xlsx in {OUTPUT_DIR})")
            continue

        cat_df = _read_sheet(path, "Catalog")
        if not cat_df.empty:
            cat_df.insert(0, "Brand", brand)
            catalog_frames.append(cat_df)
        print(f"  ✓  {brand}: read {filename}")

    # On-hand inventory now lives in a standalone workbook, not per-brand
    # Inventory Tracker sheets (split 2026-07-23). It already has its own
    # Brand column, so no insert needed here.
    inv_path, inv_filename = _latest_inventory_file()
    if inv_path:
        combined_inv = _read_sheet(inv_path, "Inventory", required_col="SKU")
        print(f"  ✓  Inventory: read {inv_filename}")
    else:
        print(f"  ⚠  No Filament_Inventory_vN.xlsx found in {OUTPUT_DIR}, "
              f"skipping inventory aggregation")
        combined_inv = pd.DataFrame()

    combined_catalog = pd.concat(catalog_frames, ignore_index=True) if catalog_frames else pd.DataFrame()

    # Drop formula-result / fill-only columns that don't carry over cleanly via pandas
    for col in ("Reorder Flag", "Color Swatch", "Total Value"):
        if col in combined_inv.columns:
            combined_inv.drop(columns=[col], inplace=True)
    if "Color Swatch" in combined_catalog.columns:
        combined_catalog.drop(columns=["Color Swatch"], inplace=True)

    return combined_catalog, combined_inv


# ── Dashboard sheet ────────────────────────────────────────────────────────────
def _build_dashboard(wb: Workbook, cat_df: pd.DataFrame, inv_df: pd.DataFrame):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F3864"

    _merge(ws, 1, 1, 1, 8, "📦  Master Filament Index  —  Dashboard",
           bg=P["hdr_bg"], size=14)
    ws.row_dimensions[1].height = 34

    r = 3
    _section(ws, r, "  CROSS-BRAND CATALOG SUMMARY", 8)

    brand_counts = cat_df["Brand"].value_counts() if not cat_df.empty else {}
    tier_counts  = cat_df["Tier"].value_counts()  if ("Tier" in cat_df.columns and not cat_df.empty) else {}

    stats = [
        ("Total Catalog Entries (all brands)",
         len(cat_df) if not cat_df.empty else 0, "FFFFFF"),
        ("Tier S entries",
         int(tier_counts.get("S", 0)), P["tier_s"]),
        ("Tier A entries",
         int(tier_counts.get("A", 0)), P["tier_a"]),
        ("Tier B entries",
         int(tier_counts.get("B", 0)), P["tier_b"]),
        ("Tier C entries",
         int(tier_counts.get("C", 0)), P["tier_c"]),
        ("Brands with AMS Adapter Required",
         int((cat_df["AMS Adapter Required"] == "Yes").sum())
         if "AMS Adapter Required" in cat_df.columns else 0, P["ams_warn"]),
    ]
    for label, val, vbg in stats:
        r += 1
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        _sc(lc, bg=P["row_hl"], bold=True)
        _sc(vc, bg=vbg, bold=True, align="center")
        ws.row_dimensions[r].height = 18

    r += 2
    _section(ws, r, "  ENTRIES PER BRAND", 8)
    r += 1
    bh = ws.cell(row=r, column=1, value="Brand")
    _sc(bh, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, align="center")
    ch = ws.cell(row=r, column=2, value="Catalog Entries")
    _sc(ch, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, align="center")
    ih = ws.cell(row=r, column=3, value="Inventory Rows")
    _sc(ih, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, align="center")
    ws.row_dimensions[r].height = 20

    inv_brand_counts = inv_df["Brand"].value_counts() if not inv_df.empty else {}
    for idx, brand in enumerate(BRANDS):
        r += 1
        alt = idx % 2 == 1
        bg = P["alt"] if alt else "FFFFFF"
        bc = ws.cell(row=r, column=1, value=brand)
        _sc(bc, bg=bg)
        cc = ws.cell(row=r, column=2, value=int(brand_counts.get(brand, 0)))
        _sc(cc, bg=bg, align="center")
        ic = ws.cell(row=r, column=3, value=int(inv_brand_counts.get(brand, 0)))
        _sc(ic, bg=bg, align="center")
        ws.row_dimensions[r].height = 18

    r += 2
    _section(ws, r, "  INVENTORY SUMMARY (ALL BRANDS)", 8)
    if not inv_df.empty:
        total_qty = int(inv_df["Quantity in Stock"].fillna(0).sum()) \
            if "Quantity in Stock" in inv_df.columns else 0
        paid_col = "Price Paid"
        if paid_col in inv_df.columns and "Quantity in Stock" in inv_df.columns:
            total_val = (inv_df[paid_col].fillna(0) *
                         inv_df["Quantity in Stock"].fillna(0)).sum()
        else:
            total_val = 0.0
    else:
        total_qty, total_val = 0, 0.0

    inv_stats = [
        ("Total Spools in Stock", total_qty,              "FFFFFF",        None),
        ("Total Inventory Value", round(total_val, 2),    "FFFFFF",        "$#,##0.00"),
    ]
    for label, val, vbg, fmt in inv_stats:
        r += 1
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        _sc(lc, bg=P["row_hl"], bold=True)
        _sc(vc, bg=vbg, bold=True, align="center")
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[r].height = 18

    r += 2
    note = ws.cell(row=r, column=1,
                   value=f"ℹ  This dashboard is generated from the {len(BRANDS)} brand files. "
                         "Re-run master_index.py to refresh after updating any brand file.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    note.font = _font(italic=True, color="595959")
    note.alignment = _align("left")
    ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    for col in "DEFGH":
        ws.column_dimensions[col].width = 10


# ── Cross-Brand Catalog sheet ──────────────────────────────────────────────────
def _build_catalog(wb: Workbook, cat_df: pd.DataFrame):
    ws = wb.create_sheet("Catalog")
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "1F3864"

    _merge(ws, 1, 1, 1, 20,
           "Master Filament Index  —  Cross-Brand Catalog  (read from brand files — edit in source)",
           bg=P["hdr_bg"], size=12)
    ws.row_dimensions[1].height = 28

    if cat_df.empty:
        ws.cell(row=3, column=1, value="No brand files found — run generate_all.py first")
        return

    # Insert Color Swatch column after Color Name
    raw_cols = list(cat_df.columns)
    hex_idx_in_df = next((i for i,c in enumerate(raw_cols) if c == "Hex Code"), None)
    color_name_idx = next((i for i,c in enumerate(raw_cols) if c == "Color Name"), None)

    # Build display column list: insert "Color Swatch" after "Color Name"
    if color_name_idx is not None:
        cols = raw_cols[:color_name_idx+1] + ["Color Swatch"] + raw_cols[color_name_idx+1:]
    else:
        cols = raw_cols
    SWATCH_COL_IDX = next((i+1 for i,c in enumerate(cols) if c == "Color Swatch"), None)
    HEX_COL_IDX    = next((i+1 for i,c in enumerate(cols) if c == "Hex Code"), None)

    _header_row(ws, 2, cols)

    TIER_COL  = next((get_column_letter(i+1) for i,c in enumerate(cols) if c == "Tier"),  None)
    AMS_COLS  = [get_column_letter(i+1) for i,c in enumerate(cols)
                 if c in ("AMS X/P", "AMS Lite", "AMS 2 Pro", "AMS HT")]
    ADAP_COL  = next((get_column_letter(i+1) for i,c in enumerate(cols)
                      if c == "AMS Adapter Required"), None)

    # Dynamic center-align + hex-fill column lookup (by header name, not hardcoded index)
    CENTER_NAMES = {"Color Swatch", "Hex Code", "Diameter", "Diameter Tolerance", "Spool Type",
                    "AMS Adapter Required", "Print Temp Range", "Bed Temp Range",
                    "Drying Temp / Time", "AMS X/P", "AMS Lite", "AMS 2 Pro",
                    "AMS HT", "Tier"}
    CENTER_COLS = {i+1 for i, c in enumerate(cols) if c in CENTER_NAMES}

    first_data = 3
    for row_idx, (_, row_data) in enumerate(cat_df.iterrows()):
        row = row_idx + first_data
        alt = row_idx % 2 == 1
        bg  = P["alt"] if alt else "FFFFFF"

        # Build row values, inserting blank for the swatch slot
        row_vals = list(row_data)
        if color_name_idx is not None:
            row_vals = row_vals[:color_name_idx+1] + [""] + row_vals[color_name_idx+1:]

        for col_idx, val in enumerate(row_vals, 1):
            display = "" if (isinstance(val, float) and pd.isna(val)) else val
            c = ws.cell(row=row, column=col_idx, value=display if col_idx != SWATCH_COL_IDX else "")
            _sc(c, bg=bg, align="center" if col_idx in CENTER_COLS else "left")

            # Color Swatch: fill from hex code
            if col_idx == SWATCH_COL_IDX:
                # Get hex from the hex column (next col)
                hex_display = row_vals[HEX_COL_IDX - 1] if HEX_COL_IDX else ""
                if isinstance(hex_display, str) and hex_display.startswith("#"):
                    hexval = hex_display.lstrip("#")
                    if len(hexval) == 6:
                        c.value = ""
                        c.fill  = _fill(hexval)
                        c.border = _border()
                        c.alignment = _align("center")

            # Hex Code chip — tinted background
            elif col_idx == HEX_COL_IDX and isinstance(display, str) and display.startswith("#"):
                hexval = display.lstrip("#")
                if len(hexval) == 6:
                    c.fill = _fill(_tint_hex(hexval, 0.12))
                    c.font = Font(name=FONT_NAME, size=9, color=_argb("444444"))

        ws.row_dimensions[row].height = 18

    last_cf = first_data + len(cat_df) + 5
    if TIER_COL:
        for val, bg, fg in (("S", P["tier_s"], P["tier_s_fg"]),
                             ("A", P["tier_a"], P["tier_a_fg"]),
                             ("B", P["tier_b"], P["tier_b_fg"]),
                             ("C", P["tier_c"], P["tier_c_fg"])):
            _cf_eq(ws, TIER_COL, first_data, last_cf, val, bg, fg)
    for col in AMS_COLS:
        _cf_has(ws, col, first_data, last_cf, "✓", P["ams_ok"],   P["ams_ok_fg"])
        _cf_has(ws, col, first_data, last_cf, "✗", P["ams_no"],   P["ams_no_fg"])
        _cf_has(ws, col, first_data, last_cf, "⚠", P["ams_warn"], P["ams_warn_fg"])
    if ADAP_COL:
        _cf_eq(ws, ADAP_COL, first_data, last_cf, "Yes", P["ams_warn"], P["ams_warn_fg"])

    # Column widths
    ws.column_dimensions["A"].width = 16  # Brand
    for i in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for i, col in enumerate(cols, 1):
        if col == "Color Swatch":
            ws.column_dimensions[get_column_letter(i)].width = 10
        elif col == "Hex Code":
            ws.column_dimensions[get_column_letter(i)].width = 10
        elif col in ("Product Name", "Tier Rationale", "Notes"):
            ws.column_dimensions[get_column_letter(i)].width = 28
        elif col in ("SKU / Part No.", "Color Name"):
            ws.column_dimensions[get_column_letter(i)].width = 18


# ── Master Inventory sheet ─────────────────────────────────────────────────────
def _build_inventory(wb: Workbook, inv_df: pd.DataFrame):
    ws = wb.create_sheet("Inventory")
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "375623"

    _merge(ws, 1, 1, 1, 20,
           "Master Filament Index  —  Inventory  (from Filament_Inventory_vN.xlsx)",
           bg=P["hdr_bg"], size=12)
    ws.row_dimensions[1].height = 28

    if inv_df.empty:
        ws.cell(row=3, column=1, value="No inventory data found — run build_inventory.py first")
        return

    cols = list(inv_df.columns)

    # Find key column positions in (cols = inv_df.columns)
    color_name_idx_inv = next((i for i,c in enumerate(cols) if c == "Color Name"), None)
    raw_hex_idx    = next((i for i,c in enumerate(cols) if c == "Hex Code"), None)

    # Insert "Color Swatch" after "Color Name"
    if color_name_idx_inv is not None:
        display_cols = cols[:color_name_idx_inv+1] + ["Color Swatch"] + cols[color_name_idx_inv+1:]
    else:
        display_cols = cols

    _header_row(ws, 2, display_cols)
    ws.row_dimensions[2].height = 34
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_cols))}{ws.max_row}"
    SWATCH_COL_IDX_INV = next((i+1 for i,c in enumerate(display_cols) if c == "Color Swatch"), None)
    HEX_COL_IDX_INV    = next((i+1 for i,c in enumerate(display_cols) if c == "Hex Code"), None)
    PRICE_COLS    = {i+1 for i,c in enumerate(display_cols) if c in ("List Price", "Price Paid")}
    QTY_COL_IDX   = next((i+1 for i,c in enumerate(display_cols) if c == "Quantity in Stock"), None)
    CENTER_NAMES  = {"Color Swatch", "Hex Code", "Diameter", "Quantity in Stock",
                     "Opened / Sealed", "Condition", "Moisture Rating"}
    CENTER_COLS   = {i+1 for i,c in enumerate(display_cols) if c in CENTER_NAMES}

    first_data = 3
    pulled_count = len(inv_df)

    # ── Pulled rows (from brand files) ──
    for row_idx, (_, row_data) in enumerate(inv_df.iterrows()):
        row = row_idx + first_data
        alt = row_idx % 2 == 1
        bg  = P["alt"] if alt else "FFFFFF"

        # Insert blank swatch slot
        row_list = list(row_data)
        if color_name_idx_inv is not None:
            row_list = row_list[:color_name_idx_inv+1] + [""] + row_list[color_name_idx_inv+1:]

        for col_idx, val in enumerate(row_list, 1):
            display = "" if (isinstance(val, float) and pd.isna(val)) else val
            c = ws.cell(row=row, column=col_idx, value="" if col_idx == SWATCH_COL_IDX_INV else display)
            _sc(c, bg=bg, align="center" if col_idx in CENTER_COLS else "left")

            if col_idx == SWATCH_COL_IDX_INV:
                hex_display = row_list[HEX_COL_IDX_INV - 1] if HEX_COL_IDX_INV else ""
                if isinstance(hex_display, str) and hex_display.startswith("#"):
                    hexval = hex_display.lstrip("#")
                    if len(hexval) == 6:
                        c.fill  = _fill(hexval)
                        c.border = _border()
                        c.alignment = _align("center")
            elif col_idx == HEX_COL_IDX_INV and isinstance(display, str) and display.startswith("#"):
                hexval = display.lstrip("#")
                if len(hexval) == 6:
                    c.fill = _fill(_tint_hex(hexval, 0.12))
                    c.font = Font(name=FONT_NAME, size=9, color=_argb("444444"))
            elif col_idx in PRICE_COLS and isinstance(display, (int, float)):
                c.number_format = "$#,##0.00"
            elif col_idx == QTY_COL_IDX and isinstance(display, (int, float)):
                c.number_format = "#,##0"
        ws.row_dimensions[row].height = 18

    # ── Manual rows section ──
    divider_row = first_data + pulled_count + 1
    ws.merge_cells(start_row=divider_row, start_column=1,
                   end_row=divider_row, end_column=len(cols))
    dc = ws.cell(row=divider_row, column=1,
                 value=f"▼  MANUAL ROWS — Add brands / spools not in the {len(BRANDS)} brand files below")
    _sc(dc, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, border=False)
    ws.row_dimensions[divider_row].height = 20

    # 30 blank manual rows with light yellow background
    for extra in range(1, 31):
        row = divider_row + extra
        for col_idx in range(1, len(display_cols) + 1):
            c = ws.cell(row=row, column=col_idx, value="")
            _sc(c, bg=P["manual"])
        ws.row_dimensions[row].height = 18

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(2, len(display_cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for i, col in enumerate(display_cols, 1):
        if col == "Color Swatch":
            ws.column_dimensions[get_column_letter(i)].width = 10
        elif col == "Hex Code":
            ws.column_dimensions[get_column_letter(i)].width = 10
        elif col in ("Product Name", "Notes"):
            ws.column_dimensions[get_column_letter(i)].width = 24
        elif col in ("SKU", "Color Name"):
            ws.column_dimensions[get_column_letter(i)].width = 18


# ── Main ───────────────────────────────────────────────────────────────────────
def build_master_index():
    print("Reading brand files...")
    cat_df, inv_df = _load_all_data()
    print(f"  Catalog:   {len(cat_df)} total rows across {cat_df['Brand'].nunique() if not cat_df.empty else 0} brands")
    print(f"  Inventory: {len(inv_df)} total rows")

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, cat_df, inv_df)
    _build_catalog(wb, cat_df)
    _build_inventory(wb, inv_df)

    pattern  = r'^master_index_11_(\d+)\.xlsx$'
    template = 'master_index_11_{n}.xlsx'
    out_path, version, prev = next_versioned_path(OUTPUT_DIR, pattern, template)
    wb.save(out_path)
    print(f"\n✓  {os.path.basename(out_path)} written to: {out_path}")
    if prev:
        print(f"   ⚠  Delete old version from the repo after uploading: {prev}")


if __name__ == "__main__":
    build_master_index()
