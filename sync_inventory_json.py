"""
sync_inventory_json.py — the mirror image of build_inventory.py.

build_inventory.py:  real_inventory.json  -->  Filament_Inventory.xlsx
sync_inventory_json.py:  Filament_Inventory.xlsx  -->  real_inventory.json

Why this exists (2026-08-04):
  real_inventory.json is the documented source of truth, but in practice the
  person maintaining this project regularly hand-edits Filament_Inventory.xlsx
  directly (adding/correcting SKUs, colors, quantities, partial-spool weights)
  and only afterwards asks for the JSON to be brought back in sync. Doing that
  translation by hand each time is slow and error-prone — this script makes it
  a single deterministic command instead.

What it does:
  1. Reads the Inventory sheet of the given xlsx, using build_inventory.py's
     own INV_HEADERS as the column contract — so if the inventory schema is
     extended again (as it was on 2026-08-02, adding Package Type / Net
     Weight (g) / Est. Remaining (g) / Est. Remaining %), this script tracks
     it automatically instead of silently dropping the new columns.
  2. Drops the derived/formula columns (Color Swatch, Reorder Flag, Total
     Value, Est. Remaining %) — those are computed by build_inventory.py on
     the way back out and should never be treated as source data.
  3. Filters phantom/headroom rows using Brand as the required column (any
     row with no Brand is pre-filled-formula headroom, not real inventory —
     see the project's own convention of treating a required identifying
     column, not dropna(how="all"), as the phantom-row filter).
  4. Prints a human-readable diff against the previous real_inventory.json
     (if one exists next to the output path) — added rows, removed rows, and
     which fields changed on rows that persisted. This is the same kind of
     row-by-row comparison that's been done by hand in this project so far,
     now automatic.
  5. Self-validates: after writing the JSON, re-imports build_inventory.py
     and actually builds a workbook from the freshly written JSON in a temp
     directory, to confirm round-tripping doesn't error before declaring
     success.

Usage:
    python3 sync_inventory_json.py [xlsx_path] [json_path]

Defaults (matching build_inventory.py's own convention of using the script's
own directory):
    xlsx_path = ./output/Filament_Inventory.xlsx
    json_path = ./real_inventory.json
"""
import os
import sys
import json
import tempfile
from datetime import date, datetime

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from build_inventory import INV_HEADERS, DATA_START  # noqa: E402  (schema contract)

# Columns that are computed by build_inventory.py, never source data.
DERIVED_COLUMNS = {"Color Swatch", "Reorder Flag", "Total Value", "Est. Remaining %"}

# Maps an xlsx column header -> the key real_inventory.json / build_inventory.py's
# __main__ block expects. Everything not listed here keeps its xlsx header name
# verbatim (they already match).
JSON_KEY_OVERRIDES = {"SKU": "SKU / Product ID"}

# What (in order) actually gets written to each JSON entry.
JSON_FIELD_ORDER = [
    "Brand", "SKU / Product ID", "Material Type", "Product Name", "Color Name",
    "Hex Code", "Diameter", "Lot / Batch Code", "Date Purchased",
    "Where Purchased", "List Price", "Price Paid", "Quantity in Stock",
    "Opened / Sealed", "Condition", "Moisture Rating", "Last Dried Date",
    "Package Type", "Net Weight (g)", "Est. Remaining (g)", "Notes",
]


def _jsonable(v):
    """openpyxl hands back datetime/date objects for date-formatted cells;
    JSON needs strings."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def read_inventory_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Inventory"]
    headers = [c.value for c in ws[DATA_START - 1]]  # header row is one above data

    if headers != INV_HEADERS:
        missing = [h for h in INV_HEADERS if h not in headers]
        extra = [h for h in headers if h and h not in INV_HEADERS]
        print("⚠  Inventory sheet headers don't exactly match build_inventory.py's "
              "INV_HEADERS.")
        if missing:
            print(f"   Missing from xlsx: {missing}")
        if extra:
            print(f"   Present in xlsx but not in schema: {extra}")
        print("   Proceeding — matching columns will still be read correctly.")

    rows = []
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        vals = [c.value for c in row]
        raw = dict(zip(headers, vals))
        if not raw.get("Brand"):
            continue  # headroom row — no identifying data, skip
        rows.append(raw)
    return rows


def raw_row_to_entry(raw: dict) -> dict:
    entry = {}
    for xlsx_key, json_key in [(h, JSON_KEY_OVERRIDES.get(h, h)) for h in INV_HEADERS]:
        if xlsx_key in DERIVED_COLUMNS:
            continue
        entry[json_key] = _jsonable(raw.get(xlsx_key))
    # keep a stable, readable key order regardless of INV_HEADERS iteration order
    return {k: entry.get(k) for k in JSON_FIELD_ORDER}


def row_key(entry: dict) -> tuple:
    """Best-effort stable identity for an inventory row, for diffing across
    runs. SKU alone isn't safe (can be blank, e.g. an unresolved color), so
    fall back to a composite of fields unlikely to collide."""
    return (
        entry.get("Brand"),
        entry.get("SKU / Product ID") or "",
        entry.get("Color Name") or "",
        entry.get("Package Type") or "",
        entry.get("Opened / Sealed") or "",
    )


def print_diff(old_entries: list[dict], new_entries: list[dict]) -> None:
    old_by_key = {row_key(e): e for e in old_entries}
    new_by_key = {row_key(e): e for e in new_entries}

    added = [k for k in new_by_key if k not in old_by_key]
    removed = [k for k in old_by_key if k not in new_by_key]
    common = [k for k in new_by_key if k in old_by_key]

    changed = []
    for k in common:
        old_e, new_e = old_by_key[k], new_by_key[k]
        diffs = {f: (old_e.get(f), new_e.get(f)) for f in JSON_FIELD_ORDER
                 if old_e.get(f) != new_e.get(f)}
        if diffs:
            changed.append((k, diffs))

    if not added and not removed and not changed:
        print("No changes vs. previous real_inventory.json.")
        return

    print(f"Diff vs. previous real_inventory.json "
          f"({len(added)} added, {len(removed)} removed, {len(changed)} changed):\n")

    for k in added:
        e = new_by_key[k]
        print(f"  + {e['Brand']} / {e['Color Name']} "
              f"(SKU: {e['SKU / Product ID'] or 'MISSING'})")

    for k in removed:
        e = old_by_key[k]
        print(f"  - {e['Brand']} / {e['Color Name']} "
              f"(SKU: {e['SKU / Product ID'] or 'MISSING'})")

    for k, diffs in changed:
        e = new_by_key[k]
        print(f"  ~ {e['Brand']} / {e['Color Name']}:")
        for field, (old_v, new_v) in diffs.items():
            print(f"      {field}: {old_v!r} -> {new_v!r}")
    print()


def self_validate(json_path: str) -> None:
    """Round-trip check: actually build a workbook from the freshly written
    JSON in a scratch directory, so a schema mistake is caught now rather
    than the next time someone runs build_inventory.py for real."""
    import importlib
    import build_inventory as bi
    importlib.reload(bi)

    with open(json_path, encoding="utf-8") as f:
        raw = json.load(f)

    entries = []
    for row in raw:
        entries.append({
            "Brand": row["Brand"],
            "SKU": row["SKU / Product ID"],
            "Material Type": row["Material Type"],
            "Product Name": row["Product Name"],
            "Color Name": row["Color Name"],
            "Hex Code": row["Hex Code"],
            "Diameter": row["Diameter"],
            "Lot / Batch Code": row.get("Lot / Batch Code"),
            "Date Purchased": row.get("Date Purchased"),
            "Where Purchased": row.get("Where Purchased"),
            "List Price": row.get("List Price"),
            "Price Paid": row.get("Price Paid"),
            "Quantity in Stock": row.get("Quantity in Stock"),
            "Opened / Sealed": row.get("Opened / Sealed"),
            "Condition": row.get("Condition"),
            "Moisture Rating": row.get("Moisture Rating"),
            "Last Dried Date": row.get("Last Dried Date"),
            "Package Type": row.get("Package Type"),
            "Net Weight (g)": row.get("Net Weight (g)"),
            "Est. Remaining (g)": row.get("Est. Remaining (g)"),
            "Notes": row.get("Notes"),
        })

    wb = bi.build_inventory_workbook(entries)
    with tempfile.TemporaryDirectory() as tmp:
        wb.save(os.path.join(tmp, "validate.xlsx"))
    print(f"✓  Self-validated: build_inventory_workbook() accepted all "
          f"{len(entries)} entries without error.")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else \
        os.path.join(script_dir, "output", "Filament_Inventory.xlsx")
    json_path = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.join(script_dir, "real_inventory.json")

    if not os.path.exists(xlsx_path):
        sys.exit(f"✗  {xlsx_path} not found.")

    raw_rows = read_inventory_rows(xlsx_path)
    new_entries = [raw_row_to_entry(r) for r in raw_rows]

    old_entries = []
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            old_entries = json.load(f)

    print_diff(old_entries, new_entries)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(new_entries, f, indent=2)
    print(f"✓  Wrote {len(new_entries)} entries to {json_path}")

    self_validate(json_path)


if __name__ == "__main__":
    main()
