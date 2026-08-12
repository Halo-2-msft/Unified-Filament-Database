"""Filament Reference System — v3 Template Engine"""

from __future__ import annotations
import os
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

MAX_DATA_ROWS = 700

# ── Palette ────────────────────────────────────────────────────────────────────
P = {
    "hdr_bg":     "1F3864", "hdr_fg":     "FFFFFF",
    "sec_bg":     "2F5496", "sec_fg":     "FFFFFF",
    "alt":        "EEF3FB",
    "cfg_bg":     "FFF2CC", "cfg_fg":     "7F6000",
    "tier_s":     "FFD700", "tier_s_fg":  "7F6000",
    "tier_a":     "92D050", "tier_a_fg":  "276221",
    "tier_b":     "BDD7EE", "tier_b_fg":  "1F3864",
    "tier_c":     "F4B8B8", "tier_c_fg":  "9C0006",
    "ams_ok":     "C6EFCE", "ams_ok_fg":  "276221",
    "ams_no":     "FFC7CE", "ams_no_fg":  "9C0006",
    "ams_warn":   "FFEB9C", "ams_warn_fg":"9C5700",
    "reorder":    "FF0000", "reord_fg":   "FFFFFF",
    "ok_bg":      "C6EFCE", "ok_fg":      "276221",
    "warn_bg":    "FFEB9C", "warn_fg":    "9C5700",
    "border":     "B8CCE4", "row_hl":     "F2F2F2",
    "disc_bg":    "E8E8E8", "disc_fg":    "888888",
    "hex_bg":     "F2F2F2",   # light chip background for hex code cells
}
FONT_NAME = "Arial"


# ── Version discovery ──────────────────────────────────────────────────────────
# Every generated workbook in this project follows a "keep only the single
# highest version" convention (Bambu_Lab_filaments_v3_10.xlsx,
# Filament_Inventory_v2.xlsx, master_index_11_8.xlsx, etc). These two helpers
# let every generator script discover the current highest version on disk and
# write the *next* one, instead of hardcoding an unversioned filename that
# drifts from what's actually in the repo.

def find_latest_version(directory: str, pattern: str):
    """Scan `directory` for filenames matching `pattern` (a regex with exactly
    one capturing group for the integer version number) and return the
    (version_number, filename) pair for the highest version found.

    Returns (None, None) if the directory doesn't exist yet or no file
    matches the pattern.
    """
    if not os.path.isdir(directory):
        return None, None
    rx = re.compile(pattern)
    best = None
    for fname in os.listdir(directory):
        m = rx.match(fname)
        if m:
            n = int(m.group(1))
            if best is None or n > best[0]:
                best = (n, fname)
    return best if best else (None, None)


def next_versioned_path(directory: str, pattern: str, template: str):
    """Given a regex `pattern` (one capture group = version number) and a
    `template` string containing '{n}' where the version number goes,
    return (full_path, version_number, previous_filename_or_None) for the
    NEXT version to write.

    Example:
        pattern  = r'^Bambu_Lab_filaments_v3_(\\d+)\\.xlsx$'
        template = 'Bambu_Lab_filaments_v3_{n}.xlsx'
    """
    latest_n, latest_fname = find_latest_version(directory, pattern)
    next_n = 1 if latest_n is None else latest_n + 1
    fname = template.format(n=next_n)
    return os.path.join(directory, fname), next_n, latest_fname


# ── Micro helpers ──────────────────────────────────────────────────────────────
def _argb(h: str) -> str:
    """Ensure full-opacity ARGB 8-char hex (openpyxl silently prepends 00 to 6-char)."""
    return h if len(h) == 8 else f"FF{h}"

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=_argb(hex6))

def _font(size=10, bold=False, color="000000", italic=False, strike=False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, color=_argb(color),
                italic=italic, strikethrough=strike)

def _border() -> Border:
    s = Side(style="thin", color=_argb(P["border"]))
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _contrasting_fg(hex6: str) -> str:
    r, g, b = int(hex6[:2],16), int(hex6[2:4],16), int(hex6[4:6],16)
    return "000000" if (0.299*r + 0.587*g + 0.114*b)/255 > 0.5 else "FFFFFF"

def _tint(hex6: str, factor: float = 0.15) -> str:
    """Blend hex6 toward white (factor=0 → white, factor=1 → original)."""
    r = int(int(hex6[:2],16)*factor + 255*(1-factor))
    g = int(int(hex6[2:4],16)*factor + 255*(1-factor))
    b = int(int(hex6[4:6],16)*factor + 255*(1-factor))
    return f"{r:02X}{g:02X}{b:02X}"

def _sc(cell, bg=None, fg="000000", bold=False, size=10,
        align="left", wrap=True, italic=False, border=True, strike=False):
    cell.font = _font(size=size, bold=bold, color=fg, italic=italic, strike=strike)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align("center" if align == "center" else "left", wrap)
    if border:
        cell.border = _border()

def _header_row(ws, row: int, headers: list, bg=None, fg=None, height=30):
    bg, fg = bg or P["hdr_bg"], fg or P["hdr_fg"]
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        _sc(c, bg=bg, fg=fg, bold=True, size=10, align="center")
    ws.row_dimensions[row].height = height

def _set_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _merge(ws, r1, c1, r2, c2, value="", bg=None, fg="FFFFFF",
           bold=True, size=12, align="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = _font(size=size, bold=bold, color=fg)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(align)

def _section(ws, row: int, text: str, ncols: int = 8):
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    _sc(c, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, border=False)
    ws.row_dimensions[row].height = 18

def _dv(formula1: str) -> DataValidation:
    return DataValidation(type="list", formula1=formula1,
                          allow_blank=True, showErrorMessage=False)

def _cf(ws, col: str, r1: int, r2: int, formula: str, bg: str, fg="000000"):
    rng = f"{col}{r1}:{col}{r2}"
    ds = DifferentialStyle(font=Font(name=FONT_NAME, color=_argb(fg)),
                           fill=PatternFill(bgColor=_argb(bg)))
    ws.conditional_formatting.add(rng, Rule(type="expression",
                                            formula=[formula], dxf=ds))

def _cf_eq(ws, col, r1, r2, val, bg, fg="000000"):
    _cf(ws, col, r1, r2, f'${col}{r1}="{val}"', bg, fg)

def _cf_has(ws, col, r1, r2, val, bg, fg="000000"):
    _cf(ws, col, r1, r2, f'NOT(ISERROR(SEARCH("{val}",${col}{r1})))', bg, fg)

def _comment(ws, cell_ref: str, text: str, width=230, height=100):
    c = ws[cell_ref]
    cm = Comment(text, "Filament Ref v3")
    cm.width, cm.height = width, height
    c.comment = cm


# ── Config sheet ───────────────────────────────────────────────────────────────
# Config!$B$3 = Reorder Threshold

def _build_config(wb: Workbook, brand: str):
    ws = wb.create_sheet("Config")
    ws.sheet_properties.tabColor = "7F7F7F"
    _merge(ws, 1, 1, 1, 3, "⚙  Configuration", bg=P["hdr_bg"], size=12)
    ws.row_dimensions[1].height = 26
    _header_row(ws, 2, ["Setting", "Value", "Notes"], bg=P["sec_bg"], height=22)

    sku_note = {
        "Bambu Lab":      "BL-XXXX reference codes — Bambu does not publish clean SKUs",
        "Polymaker":      "PolyLite PLA: catalog codes PM-PLA-01–29, official hex confirmed at shop.polymaker.com. Other lines: reference codes",
        "Prusament":      "Reference codes (PRUSPL-xxx) — verify real codes at prusa3d.com",
        "SUNLU":          "Reference codes (SL-xxx) — verify ASINs on Amazon",
        "eSUN":           "Reference codes based on eSUN naming convention",
        "Overture":       "Catalog codes OV-{Material}-{Color} (redesigned 2026-07 to fix SKU collisions) — reference codes, verify ASINs on Amazon",
        "Hatchbox":       "Reference codes (HX-xxx) — verify ASINs on Amazon",
        "Creality Hyper": "Reference codes (CRH-xxx) — verify ASINs on Amazon",
        "MatterHackers":  "Reference codes (MH-xxx) — verify at matterhackers.com",
    }.get(brand, "Reference codes — verify at manufacturer site")

    settings = [
        ("Brand",       brand, ""),
        ("File Version","v3",  ""),
        ("Generated",   date.today().isoformat(),
         "Date this file was last generated by generate_all.py"),
        ("SKU Note",    sku_note,
         "SKU column: confirmed mfr codes noted above; others are internal reference codes"),
        ("Hex Note",    "Per-row confirmation status is tracked in each Catalog row's Notes column "
                        "(manufacturer TDS-confirmed, official product-page confirmed, or unconfirmed placeholder). "
                        "#999999 in the Hex Code column always means unconfirmed placeholder, not a real color.",
         "Approximate hex = visually representative; not manufacturer-confirmed"),
        ("Inventory Note", "This workbook covers catalog reference data only. On-hand spool inventory "
                           "is tracked separately in Filament_Inventory.xlsx.",
         "Split from per-brand Inventory Tracker sheets, 2026-07-23"),
    ]
    for i, (label, val, note) in enumerate(settings, 3):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=3, value=note)
        _sc(ws.cell(row=i, column=1), bold=True)
        _sc(c, bg=P["cfg_bg"], fg=P["cfg_fg"], bold=(i == 3), align="center")
        _sc(ws.cell(row=i, column=3), italic=True)
        ws.row_dimensions[i].height = 18
    _set_widths(ws, [20, 56, 70])


# ── Catalog sheet ──────────────────────────────────────────────────────────────
# Column layout (20 cols):
# A  Material Type        B  SKU / Part No.       C  Product Name
# D  Color Name           E  Color Swatch          F  Hex Code   ← new
# G  Diameter             H  Diameter Tolerance    I  Spool Type
# J  AMS Adapter Required K  Print Temp Range      L  Bed Temp Range
# M  Drying Temp / Time   N  AMS X/P               O  AMS Lite
# P  AMS 2 Pro            Q  AMS HT                R  Tier
# S  Tier Rationale       T  Notes

CAT_HEADERS = [
    "Material Type", "SKU / Part No.", "Product Name", "Color Name",
    "Color Swatch", "Hex Code",
    "Diameter", "Diameter Tolerance", "Spool Type", "AMS Adapter Required",
    "Print Temp Range", "Bed Temp Range", "Drying Temp / Time",
    "AMS X/P", "AMS Lite", "AMS 2 Pro", "AMS HT",
    "Tier", "Tier Rationale", "Notes",
]
CAT_COL = {h: get_column_letter(i + 1) for i, h in enumerate(CAT_HEADERS)}

CAT_WIDTHS = [
    15, 18, 22, 18, 10, 10,          # A-F
    9, 16, 12, 18,                   # G-J
    15, 14, 16,                      # K-M
    9, 9, 9, 9,                      # N-Q  (AMS group — collapsible)
    9, 48, 48,                       # R-T
]

_DATA_START = 3
_LAST_CAT   = _DATA_START + MAX_DATA_ROWS


def _build_catalog(wb: Workbook, brand_data: dict):
    ws = wb.create_sheet("Catalog")
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "1F3864"

    brand   = brand_data["brand"]
    catalog = brand_data.get("catalog", [])

    _merge(ws, 1, 1, 1, len(CAT_HEADERS),
           f"{brand}  —  Filament Catalog", bg=P["hdr_bg"], size=13)
    ws.row_dimensions[1].height = 30
    _header_row(ws, 2, CAT_HEADERS)
    ws.row_dimensions[2].height = 34

    # AutoFilter on header row
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(len(CAT_HEADERS))}{_LAST_CAT}"
    )

    # CENTER_COLS: swatch(5), hex(6), diameter(7) through tier(18)
    CENTER_COLS = set(range(5, 19))   # 5-18 inclusive

    for idx, entry in enumerate(catalog):
        row     = idx + _DATA_START
        alt     = idx % 2 == 1
        is_disc = entry.get("discontinued", False)
        base_bg = P["disc_bg"] if is_disc else (P["alt"] if alt else "FFFFFF")
        text_fg = P["disc_fg"] if is_disc else "000000"
        hex_val = entry.get("color_hex", "")
        notes_val = entry.get("notes") or ""
        if is_disc and "[DISCONTINUED]" not in notes_val:
            notes_val = (f"[DISCONTINUED] {notes_val}").strip()

        vals = [
            entry.get("material_type", ""),           # A
            entry.get("sku", ""),                     # B
            entry.get("product_name", ""),            # C
            entry.get("color_name", ""),              # D
            "",                                       # E  swatch (styled below)
            f"#{hex_val}" if hex_val else "",         # F  hex code
            entry.get("diameter", "1.75mm"),          # G
            entry.get("diameter_tolerance", "N/A"),   # H
            entry.get("spool_type", "Plastic"),       # I
            entry.get("ams_adapter", "No"),           # J
            entry.get("print_temp", ""),              # K
            entry.get("bed_temp", ""),                # L
            entry.get("drying", ""),                  # M
            entry.get("ams_xp",   "✓"),               # N
            entry.get("ams_lite", "✓"),               # O
            entry.get("ams_2pro", "✓"),               # P
            entry.get("ams_ht",   "✓"),               # Q
            entry.get("tier", ""),                    # R
            entry.get("tier_rationale", ""),          # S
            notes_val,                                # T
        ]

        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            _sc(c, bg=base_bg, fg=text_fg, strike=is_disc,
                align="center" if col_idx in CENTER_COLS else "left")

        # Color swatch (col E = 5) — filled cell, no text
        if hex_val and not is_disc:
            sw = ws.cell(row=row, column=5, value="")
            sw.fill      = _fill(hex_val)
            sw.border    = _border()
            sw.alignment = _align("center")

        # Hex code cell (col F = 6) — tinted chip background
        if hex_val and not is_disc:
            hc = ws.cell(row=row, column=6, value=f"#{hex_val}")
            hc.fill      = _fill(_tint(hex_val, 0.12))
            hc.font      = _font(size=9, color="444444")
            hc.border    = _border()
            hc.alignment = _align("center")
        elif is_disc:
            pass  # already styled gray from vals loop

        ws.row_dimensions[row].height = 20

    # ── Data validations ──────────────────────────────────────────────────────
    last_dv = _DATA_START + len(catalog) + 100
    for col_name, formula in [
        ("Tier",                 '"S,A,B,C"'),
        ("Spool Type",           '"Cardboard,Plastic"'),
        ("AMS Adapter Required", '"Yes,No"'),
        ("Diameter",             '"1.75mm,2.85mm"'),
    ]:
        dv  = _dv(formula)
        col = CAT_COL[col_name]
        dv.sqref = f"{col}{_DATA_START}:{col}{last_dv}"
        ws.add_data_validation(dv)

    for col_name in ("AMS X/P", "AMS Lite", "AMS 2 Pro", "AMS HT"):
        dv  = _dv('"✓,✗,⚠"')
        col = CAT_COL[col_name]
        dv.sqref = f"{col}{_DATA_START}:{col}{last_dv}"
        ws.add_data_validation(dv)

    # ── Conditional formatting ────────────────────────────────────────────────
    tier_col = CAT_COL["Tier"]           # R
    for val, bg, fg in (
        ("S", P["tier_s"], P["tier_s_fg"]),
        ("A", P["tier_a"], P["tier_a_fg"]),
        ("B", P["tier_b"], P["tier_b_fg"]),
        ("C", P["tier_c"], P["tier_c_fg"]),
    ):
        _cf_eq(ws, tier_col, _DATA_START, _LAST_CAT, val, bg, fg)

    for col_name in ("AMS X/P", "AMS Lite", "AMS 2 Pro", "AMS HT"):
        col = CAT_COL[col_name]
        _cf_has(ws, col, _DATA_START, _LAST_CAT, "✓", P["ams_ok"],   P["ams_ok_fg"])
        _cf_has(ws, col, _DATA_START, _LAST_CAT, "✗", P["ams_no"],   P["ams_no_fg"])
        _cf_has(ws, col, _DATA_START, _LAST_CAT, "⚠", P["ams_warn"], P["ams_warn_fg"])

    adp = CAT_COL["AMS Adapter Required"]   # J
    _cf_eq(ws, adp, _DATA_START, _LAST_CAT, "Yes", P["ams_warn"], P["ams_warn_fg"])

    # Notes column — amber when abrasive/hardened nozzle required
    notes_col = CAT_COL["Notes"]             # T
    _cf_has(ws, notes_col, _DATA_START, _LAST_CAT,
            "hardened nozzle", "FFCC66", "7F4F00")

    # ── AMS column group (N–Q) — collapsible ─────────────────────────────────
    for col_letter in ("N", "O", "P", "Q"):
        ws.column_dimensions[col_letter].outlineLevel = 1

    # ── Header comments ───────────────────────────────────────────────────────
    _comment(ws, "B2",
             "SKU / Part No.\n\nPolymaker PolyLite PLA: confirmed manufacturer codes "
             "(PA02001–PA02029).\nAll other brands: internal reference codes.\n"
             "See Config sheet → SKU Note for details.")
    _comment(ws, "E2",
             "Color Swatch\n\nCell fill = filament color. No text — see Hex Code (col F).\n\n"
             "Official hex sources:\n"
             "• Bambu PLA Basic/Matte/PETG Basic/ABS/PLA Translucent: Bambu hex PDFs\n"
             "• Polymaker PolyLite PLA: confirmed from shop.polymaker.com\n"
             "• All other brands: approximate")
    _comment(ws, "F2",
             "Hex Code\n\nThe hex color code for this filament.\n\n"
             "Background tint is a 12% blend of the swatch color toward white,\n"
             "giving a visual connection to the adjacent swatch cell.\n\n"
             "Official hex sources: see Color Swatch comment (col E).")
    _comment(ws, "N2",
             "AMS X/P Compatibility\n\n"
             "✓  Compatible with standard AMS X/P unit\n"
             "✗  Not compatible — use external spool\n"
             "⚠  Compatible with conditions (adapter ring, special mode, etc.)\n\n"
             "Printers: X1C, X1E, P1S, P1P")
    _comment(ws, "O2",
             "AMS Lite Compatibility\n\n"
             "✓  Compatible with AMS Lite unit\n"
             "✗  Not compatible (open-frame; no ABS/ASA/PA support)\n"
             "⚠  Compatible with conditions\n\n"
             "Printers: A1, A1 mini")
    _comment(ws, "P2",
             "AMS 2 Pro Compatibility\n\n"
             "✓  Compatible with AMS 2 Pro unit\n"
             "✗  Not compatible\n"
             "⚠  Compatible with conditions\n\n"
             "Printers: H2D")
    _comment(ws, "Q2",
             "AMS HT Compatibility\n\n"
             "✓  Compatible with AMS HT unit\n"
             "✗  Not compatible\n"
             "⚠  Compatible with conditions\n\n"
             "Printers: H2D\nRequired for TPU 95A multi-color printing")

    _set_widths(ws, CAT_WIDTHS)


# ── Material Guide sheet ───────────────────────────────────────────────────────
GUIDE_HEADERS = [
    "Material", "Print Temp Range", "Bed Temp Range",
    "Enclosure Required", "AMS Compatible", "Drying Required",
    "Drying Temp / Time", "Notes",
]
GUIDE_WIDTHS = [20, 16, 14, 18, 14, 16, 16, 58]


def _build_material_guide(wb: Workbook, brand_data: dict):
    ws = wb.create_sheet("Material Guide")
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "833C00"

    brand = brand_data["brand"]
    _merge(ws, 1, 1, 1, len(GUIDE_HEADERS),
           f"{brand}  —  Material Guide", bg=P["hdr_bg"], size=13)
    ws.row_dimensions[1].height = 30
    _header_row(ws, 2, GUIDE_HEADERS)
    ws.row_dimensions[2].height = 34
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(len(GUIDE_HEADERS))}{_DATA_START + 60}"
    )

    guide = brand_data.get("material_guide", [])
    CENTER_COLS = {2, 3, 4, 5, 6, 7}

    for idx, entry in enumerate(guide):
        row     = idx + _DATA_START
        alt     = idx % 2 == 1
        base_bg = P["alt"] if alt else "FFFFFF"
        vals = [
            entry.get("material", ""),
            entry.get("print_temp", ""),
            entry.get("bed_temp", ""),
            entry.get("enclosure", "No"),
            entry.get("ams_compat", "Yes"),
            entry.get("drying_required", "No"),
            entry.get("drying", ""),
            entry.get("notes", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            _sc(c, bg=base_bg,
                align="center" if col_idx in CENTER_COLS else "left")
        ws.row_dimensions[row].height = 20

    last_g = _DATA_START + len(guide) + 10
    _cf_eq(ws, "D", _DATA_START, last_g, "Yes",  P["warn_bg"], P["warn_fg"])
    _cf_eq(ws, "E", _DATA_START, last_g, "No",   P["ams_no"],  P["ams_no_fg"])
    _cf_has(ws, "E", _DATA_START, last_g, "⚠",   P["ams_warn"], P["ams_warn_fg"])

    _set_widths(ws, GUIDE_WIDTHS)


# ── Dashboard sheet ────────────────────────────────────────────────────────────
# Catalog column refs (updated for 20-col schema):
#   C  = Product Name (COUNTA)     J  = AMS Adapter Required
#   N-Q = AMS X/P through AMS HT   R  = Tier
#   T  = Notes                     A  = Material Type

_DASH_W = 8


def _kv(ws, row: int, label: str, value, label_bg=None,
        val_bg="FFFFFF", val_format=None, bold_val=True):
    lc = ws.cell(row=row, column=1, value=label)
    vc = ws.cell(row=row, column=2, value=value)
    _sc(lc, bg=label_bg or P["row_hl"], bold=True)
    _sc(vc, bg=val_bg, bold=bold_val, align="center")
    if val_format:
        vc.number_format = val_format
    ws.row_dimensions[row].height = 18


def _build_dashboard(wb: Workbook, brand_data: dict):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "375623"

    brand   = brand_data["brand"]
    catalog = brand_data.get("catalog", [])

    # Unique ordered material types (excluding discontinued)
    mat_types, seen = [], set()
    for e in catalog:
        m = e.get("material_type", "")
        if m and m not in seen and not e.get("discontinued"):
            mat_types.append(m)
            seen.add(m)

    # ── Title ──
    r = 1
    _merge(ws, r, 1, r, _DASH_W,
           f"📦  {brand}  —  Filament Reference  |  v3",
           bg=P["hdr_bg"], size=14)
    ws.row_dimensions[r].height = 36

    # ── Catalog Summary ──
    r = 3
    _section(ws, r, "  CATALOG SUMMARY", _DASH_W)
    r += 1; _kv(ws, r, "Total Catalog Entries (active)",
                '=COUNTA(Catalog!C3:C1000)'
                '-COUNTIF(Catalog!T3:T1000,"~[DISCONTINUED~]*")')
    r += 1; _kv(ws, r, "Tier S — Premium",
                '=COUNTIF(Catalog!R3:R1000,"S")', val_bg=P["tier_s"])
    r += 1; _kv(ws, r, "Tier A — Excellent",
                '=COUNTIF(Catalog!R3:R1000,"A")', val_bg=P["tier_a"])
    r += 1; _kv(ws, r, "Tier B — Good",
                '=COUNTIF(Catalog!R3:R1000,"B")', val_bg=P["tier_b"])
    r += 1; _kv(ws, r, "Tier C — Specialty",
                '=COUNTIF(Catalog!R3:R1000,"C")', val_bg=P["tier_c"])
    r += 1; _kv(ws, r, "Require AMS Adapter Ring",
                '=COUNTIF(Catalog!J3:J1000,"Yes")', val_bg=P["warn_bg"])

    # ── Tier Legend ──
    r += 2
    _section(ws, r, "  TIER LEGEND", _DASH_W)
    for tier, bg, fg, desc in (
        ("S", P["tier_s"], P["tier_s_fg"],
         "Premium — first-party or best-in-class; optimized profiles, top AMS reliability"),
        ("A", P["tier_a"], P["tier_a_fg"],
         "Excellent — high quality, tested profiles, reliable AMS performance"),
        ("B", P["tier_b"], P["tier_b_fg"],
         "Good — solid performer with minor caveats (e.g. needs adapter ring, temp adjustment)"),
        ("C", P["tier_c"], P["tier_c_fg"],
         "Specialty — limited AMS use, requires special setup, or niche material"),
    ):
        r += 1
        tc = ws.cell(row=r, column=1, value=f"  {tier}")
        _sc(tc, bg=bg, fg=fg, bold=True, align="center")
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=_DASH_W)
        dc = ws.cell(row=r, column=2, value=desc)
        _sc(dc, border=False)
        ws.row_dimensions[r].height = 18

    # ── Inventory Note ──
    r += 2
    _section(ws, r, "  INVENTORY", _DASH_W)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_DASH_W)
    note = ws.cell(row=r, column=1,
                    value="On-hand spool inventory is tracked separately in Filament_Inventory.xlsx, not in this workbook.")
    note.font = _font(italic=True, size=10)
    note.alignment = _align("left")
    r += 1; _kv(ws, r, "File Generated",
                "=Config!$B$5", val_bg="FFFFFF", bold_val=False)

    # ── AMS Adapter Ring Status ──
    r += 2
    _section(ws, r, "  AMS ADAPTER RING STATUS", _DASH_W)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_DASH_W)
    c = ws.cell(
        row=r, column=1,
        value=('=IF(COUNTIF(Catalog!J3:J1000,"Yes")>0,'
               '"⚠  One or more filaments require the AMS Adapter Ring — see Catalog column J",'
               '"✓  No filaments in this catalog require the AMS Adapter Ring")'),
    )
    c.font      = _font(bold=True, size=10)
    c.alignment = _align("left")
    ds_warn = DifferentialStyle(
        font=Font(name=FONT_NAME, bold=True, color=_argb(P["warn_fg"])),
        fill=PatternFill(bgColor=_argb(P["warn_bg"])))
    ds_ok = DifferentialStyle(
        font=Font(name=FONT_NAME, bold=True, color=_argb(P["ok_fg"])),
        fill=PatternFill(bgColor=_argb(P["ok_bg"])))
    ws.conditional_formatting.add(
        f"A{r}:H{r}",
        Rule(type="expression", formula=[f'NOT(ISERROR(SEARCH("⚠",A{r})))'], dxf=ds_warn))
    ws.conditional_formatting.add(
        f"A{r}:H{r}",
        Rule(type="expression", formula=[f'NOT(ISERROR(SEARCH("✓",A{r})))'], dxf=ds_ok))
    ws.row_dimensions[r].height = 20

    # ── By Material Type ──
    r += 2
    _section(ws, r, "  BY MATERIAL TYPE", _DASH_W)
    r += 1
    for hdr, col in (
        ("Material Type",    1),
        ("Catalog Entries",  2),
    ):
        h = ws.cell(row=r, column=col, value=hdr)
        _sc(h, bg=P["sec_bg"], fg=P["sec_fg"], bold=True, align="center")
    ws.row_dimensions[r].height = 22

    for mat in mat_types:
        r += 1
        mc = ws.cell(row=r, column=1, value=mat)
        _sc(mc)
        cc = ws.cell(row=r, column=2,
                     value=f'=COUNTIF(Catalog!A3:A700,"{mat}")')
        _sc(cc, align="center")
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 10


# ── Public API ─────────────────────────────────────────────────────────────────
def build_workbook(brand_data: dict) -> Workbook:
    """Build a complete v3 filament workbook from brand_data."""

    raw_cat = brand_data.get("catalog", [])
    catalog_sorted = sorted(
        raw_cat,
        key=lambda e: (
            1 if e.get("discontinued") else 0,
            e.get("material_type", ""),
            e.get("color_name", ""),
        ),
    )
    brand_data = {**brand_data, "catalog": catalog_sorted}

    wb = Workbook()
    wb.remove(wb.active)
    _build_dashboard(wb, brand_data)
    _build_catalog(wb, brand_data)
    _build_material_guide(wb, brand_data)
    _build_config(wb, brand_data["brand"])
    return wb
